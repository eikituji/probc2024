import openpyxl
import os
import MySQLdb

path = os.path.dirname(__file__)
wb = openpyxl.load_workbook(f"{path}\style05.xlsx")
pat = ["0","1","2","3","4","5","6","7","8","9","０","１","２","３","４","５","６","７","８","９"]

with MySQLdb.connect(host="localhost", user="probc2024", passwd="probc2024", db="probc2024") as con:
    with con.cursor() as cur:
        sql = "SELECT 拾得物分類.物品名, 拾得物.特徴, 拾得物.色, 拾得物管理状況.変更日時, 拾得物.拾得場所, "
        sql += "拾得物.ID AS 拾得物ID, 拾得物.画像, "
        sql += "拾得物管理状況.ID AS 拾得物管理状況ID, 拾得物管理状況.変更内容 "
        sql += "FROM 拾得物管理状況 INNER JOIN "
        sql += "(SELECT 拾得物ID,MAX(変更日時) AS 日時 FROM 拾得物管理状況 GROUP BY 拾得物ID) AS 最終更新 "
        sql += "ON 拾得物管理状況.拾得物ID=最終更新.拾得物ID AND 拾得物管理状況.変更日時=最終更新.日時 "
        sql += "INNER JOIN ユーザ ON 拾得物管理状況.ユーザID=ユーザ.ID "
        sql += "INNER JOIN 所属 ON ユーザ.所属ID=所属.ID "
        sql += "INNER JOIN 拾得物 ON 拾得物管理状況.拾得物ID=拾得物.ID "
        sql += "INNER JOIN 拾得物分類 ON 拾得物.拾得物分類ID=拾得物分類.ID "
        sql += "WHERE 拾得物管理状況.変更内容 != '返還'"
        sql += "ORDER BY 拾得物管理状況.変更日時 ASC;"
        cur.execute(sql)
        rows = cur.fetchall()
        i,p = 0,0
        for row in rows:
            if (i + 10) > 46 or i == 0:
                i,p = 0,p+1
                ws = wb.copy_worksheet(wb["template"])
                ws.title = f"拾得物リスト{p}"
            ws.cell(10+i,9).value = row[0]
            if row[0] == "現金":
                j = 0
                for s in row[1][::-1]:
                    if not(s in pat): continue
                    ws.cell(10+i,8-j).value = s
                    print(s)
                    j += 1
                ws.cell(10+i,8-j).value = "¥"
            else:
                ws.cell(10+i,11).value = row[1] + row[2]
            ws.cell(10+i,17).value = row[3].strftime('%m/%d %H:%M')
            ws.cell(10+i,19).value = row[4]
            i += 4
        ws = wb.remove(wb["template"])


wb.save(f"{path}\out.xlsx")


